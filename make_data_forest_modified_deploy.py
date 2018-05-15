import matplotlib.pyplot as plt
import openpyxl as px
import numpy as np
import cPickle
import argparse
FoldNum = 4
parser = argparse.ArgumentParser()
parser.add_argument('--data_dir', default='./data/', help='path to directory with data')
opt = parser.parse_args()

output = opt.data_dir+'processed_forest_trial_new_deploy.pkl'
w = px.load_workbook(opt.data_dir+'/complete_modified_deploy.xlsx')
w1 = px.load_workbook(opt.data_dir+'new.xlsx')
p = w.get_sheet_by_name('Sheet1')
p1 = w1.get_sheet_by_name('Sheet1')

row_num = 0
PositiveData = {}
NegativeData = {}
UnknownData = {}
deployNegative = {}
deployPositive = {}

Invalid_ID = []

ID_coordinate = {}
ID_isroad = {}
ID_elevation = {}
ID_slope = {}

"""
Forest Data Format: (Column)
ID: 0
X: 1
Y: 2
IsStream:3
distStream:4
IsRoad: 5
elevation: 15
slope: 16
Features: 4,6 ~ 16
deploy_label: 17
deploy_patrol: 18
if_patrol: 19 ~ 22
if_poaching: (61,62,63, 67,68,69, 71,72,73, 76,77,78)+=4
"""
feature_id_set = set()
for (row1, row) in zip(p1.iter_rows(),p.iter_rows()):
    # # To see the column index for each category
    # if row_num == 0:
    #     line = []
    #     for k in row:
    #         line.append(k.internal_value)
    #     for i, j in enumerate(line):
    #         print i, j
    #     exit(0)
    if row_num > 1:
        integrity = 1
        line = []
        feature_id = 0
        feature_id1= 0
        for k in row:
            feature_id += 1
            try:
                if feature_id == 6:
                    line.append(float(row2[1].internal_value))
                elif feature_id == 10:
                    line.append(float(row2[4].internal_value))
                elif feature_id == 16:
                    line.append(float(row2[5].internal_value))
                elif feature_id == 17:
                    line.append(float(row2[6].internal_value))
                else:
                    line.append(float(k.internal_value))
            except TypeError:
                integrity = 0
                feature_id_set.add(feature_id)
                print row_num
                break
                # Somehow some line has None value, just skip them
            except UnicodeEncodeError:
                # Somehow the data has chinese value
                line.append(1.)
            except ValueError:
                # the tool type can be multiple
                line.append(1.)
        for k in row2:
            feature_id +=1
            feature_id1 +=1
            try:
                if feature_id1 in [2,5,6,7]:
                    continue
                elif feature_id1 == 8:
                    one_hot = np.eye(11)[k.internal_value]
                    for i in range(10):
                        line.append(float(one_hot[i]))
                elif feature_id1 in [3,4]:
                    line.append(float(k.internal_value))
            except TypeError:
                integrity = 0
                feature_id_set.add(feature_id)
                print row_num
            except UnicodeEncodeError:
                line.append(1.)
            except ValueError:
                line.append(1.)

        assert len(line) >= 4
        ID_coordinate[int(line[0])] = (int(line[1]), int(line[2]))
        ID_isroad[int(line[0])] = int(line[5])

        if integrity:
            ID_elevation[int(line[0])] = line[15]
            ID_slope[int(line[0])] = line[16]
            if line[18]>0:
                if line[17]:
                    deployPositive[int(line[0])] = line[4:5] + line[6:17] + [line[19]] + line[83:-1]
                else:
                    deployNegative[int(line[0])] = line[4:5] + line[6:17] + [line[19]] + line[83:-1]

            patrol_length = np.sum(line[19:23])
            if patrol_length > 0:
                if np.sum(line[65:68]) > 0:
                    PositiveData[300000+int(line[0])] = line[4:5] + line[6:17] + [line[20]] + line[83:-1]
                else:
                    NegativeData[300000+int(line[0])] = line[4:5] + line[6:17] + [line[20]] + line[83:-1]

                if np.sum(line[71:74]) > 0:
                    PositiveData[200000+int(line[0])] = line[4:5] + line[6:17] + [line[21]] + line[83:-1]
                else:
                    NegativeData[200000+int(line[0])] = line[4:5] + line[6:17] + [line[21]] + line[83:-1]
                if np.sum(line[75:78]) > 0:
                    PositiveData[100000+int(line[0])] = line[4:5] + line[6:17] + [line[22]] + line[83:-1]
                else:
                    NegativeData[100000+int(line[0])] = line[4:5] + line[6:17] + [line[22]] + line[83:-1]
                #if np.sum(line[77:80]) > 0:
                #    PositiveData[400000+int(line[0])] = line[4:5] + line[6:17] + [0] + line[81:-1]
                #else:
                #    NegativeData[400000+int(line[0])] = line[4:5] + line[6:17] + [0] + line[81:-1]
            else:
                UnknownData[int(line[0])] = line[4:5] + line[6:17] + [0] + line[83:-1]
        else:
            Invalid_ID.append(int(line[0]))
    row_num += 1
    row2 = row1

cPickle.dump([deployPositive, deployNegative, PositiveData, NegativeData, UnknownData, Invalid_ID, ID_coordinate, ID_isroad, ID_elevation, ID_slope],
             open(output, 'w'))

print 'Labeled data number:', len(PositiveData) + len(NegativeData)
print 'Labeled positive data:', len(PositiveData)
print 'Labeled negative data:', len(NegativeData)
print 'Unlabeled data:', len(UnknownData)
print 'Invalid data:', len(Invalid_ID)
print 'Saved data to %s' % output
print feature_id_set

"""
Labeled data number: 9569
Labeled positive data: 119
Labeled negative data: 9450
Unlabeled data: 40423
Invalid data: 208
"""
